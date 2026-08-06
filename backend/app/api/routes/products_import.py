import csv
import io
from datetime import datetime, date

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app.core.database import get_db
from app.models.catalog import Product, ProductPricing, Category, Brand, Manufacturer, Marketer
from app.models.enums import PricingChannel
from app.models.user import User
from app.api.deps import require_admin
from app.services.invoice_extraction import extract_invoice, InvoiceExtractionError
import secrets

router = APIRouter(prefix="/api/admin/products", tags=["products-import"])

# Spreadsheet-friendly identifiers, not raw DB ids — an admin filling this in
# knows category/manufacturer *names*, not their internal ids. category_slug
# and brand_slug already existed; manufacturer_name/marketer_name are looked
# up-or-created the same way ImageUploader's quick-add works in the admin UI.
# variant_group_key is a free-text key the admin makes up themselves — any
# rows sharing the same non-empty key get linked into one variant group,
# which is the natural way to express "these are the same product, different
# sizes" in a spreadsheet without needing to know product ids in advance.
TEMPLATE_HEADERS = [
    "sku", "name", "description", "image_urls", "category_slug", "brand_slug",
    "is_prescription_required",
    "manufacturer_name", "marketer_name", "country_of_origin",
    "compliance_expiry_month", "compliance_expiry_year",
    "variant_group_key", "variant_label",
    "b2c_price", "b2c_mrp", "b2c_stock", "b2c_reorder_level",
    "b2b_normal_price", "b2b_normal_mrp", "b2b_normal_min_qty", "b2b_normal_stock", "b2b_normal_reorder_level",
    "b2b_advance_price", "b2b_advance_mrp", "b2b_advance_min_qty", "b2b_advance_stock", "b2b_advance_reorder_level",
    "cnf_price", "cnf_mrp", "cnf_min_qty", "cnf_stock", "cnf_reorder_level",
    "batch_number", "batch_expiry_date", "rack_place",
    "is_spotlighted", "spotlight_order",
]

EXAMPLE_ROWS = [
    [
        "DELETE-THIS-ROW-1", "EXAMPLE ROW — DELETE BEFORE IMPORTING", "This row is a formatting example only. Delete it.", "", "diabetes-care", "accu-chek",
        "FALSE",
        "Accu-Chek Manufacturing", "", "Germany",
        "", "",
        "", "",
        999, 1200, 50, 10,
        250, 300, 15, 500, 20,
        200, 260, 30, 500, 30,
        180, 220, 10, 200, 20,
        "BATCH-2026-07", "2027-12-31", "Rack A-3",
        "FALSE", 0,
    ],
    [
        "DELETE-THIS-ROW-2", "EXAMPLE ROW — DELETE BEFORE IMPORTING", "This row shows how variant_group_key links two sizes. Delete it.", "", "medicines", "",
        "TRUE",
        "Sun Pharmaceutical Industries Ltd", "Sun Pharmaceutical Industries Ltd", "India",
        "5", "2027",
        "EXAMPLE-VARIANT-KEY", "10 Capsules",
        125, 150, 25, 10,
        "", "", "", "", "",
        "", "", "", "", "",
        "", "", "", "", "",
        "", "", "RACK-A2",
        "FALSE", 0,
    ],
]

# Real product names/SKUs deliberately never appear in this file — an
# earlier version of this template used realistic-looking example data
# (e.g. a name close to a real product), and an admin who forgot to delete
# the example rows before importing ended up with duplicate products,
# because the fake SKU didn't match anything real so it was created as new.
# Every example SKU/name here is now unmistakably a placeholder, on purpose.

# Note: this template does NOT cover product_relations (Frequently Bought
# Together / Similar / Also Bought) — those are curated per-product links
# between two existing rows, which needs the products to already exist and
# doesn't map cleanly onto a flat spreadsheet row. Use the "Related products"
# section in each product's edit screen for those.


def _build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Products"
    ws.append(TEMPLATE_HEADERS)
    for row in EXAMPLE_ROWS:
        ws.append(row)
    for col_idx, header in enumerate(TEMPLATE_HEADERS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = max(16, len(header) + 2)
    return wb


@router.get("/import-template")
def download_template(format: str = Query("xlsx", pattern="^(xlsx|csv)$"), _admin: User = Depends(require_admin)):
    """A ready-to-fill template with the exact columns the import endpoint
    expects, plus two example rows — one standalone product, one showing how
    variant_group_key links two rows together."""
    if format == "csv":
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        writer.writerow(TEMPLATE_HEADERS)
        writer.writerows(EXAMPLE_ROWS)
        byte_buffer = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))  # BOM so Excel opens UTF-8 CSVs cleanly
        return StreamingResponse(
            byte_buffer,
            media_type="text/csv",
            headers={"Content-Disposition": "attachment; filename=healthycian_product_import_template.csv"},
        )

    wb = _build_workbook()
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=healthycian_product_import_template.xlsx"},
    )


def _slugify(text: str) -> str:
    import re
    return re.sub(r"(^-|-$)", "", re.sub(r"[^a-z0-9]+", "-", text.lower().strip()))


def _parse_date(value) -> date | None:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value).strip(), "%Y-%m-%d").date()


def _parse_num(value):
    if value in (None, ""):
        return None
    return float(value)


def _parse_bool(value) -> bool:
    if isinstance(value, bool):
        return value
    return str(value).strip().lower() in ("true", "1", "yes", "y")


def _rows_from_xlsx(contents: bytes) -> list[list]:
    try:
        wb = load_workbook(io.BytesIO(contents), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="Couldn't read that file — is it a valid .xlsx?")
    ws = wb.active
    return list(ws.iter_rows(values_only=True))


def _rows_from_csv(contents: bytes) -> list[list]:
    try:
        text = contents.decode("utf-8-sig")  # tolerates a BOM if the file has one, harmless if not
    except UnicodeDecodeError:
        raise HTTPException(status_code=400, detail="Couldn't read that file as UTF-8 CSV")
    reader = csv.reader(io.StringIO(text))
    return [row for row in reader]


@router.post("/import")
async def import_products(file: UploadFile = File(...), db: Session = Depends(get_db), _admin: User = Depends(require_admin)):
    filename = (file.filename or "").lower()
    contents = await file.read()

    if filename.endswith((".xlsx", ".xlsm")):
        rows = _rows_from_xlsx(contents)
    elif filename.endswith(".csv"):
        rows = _rows_from_csv(contents)
    else:
        raise HTTPException(status_code=400, detail="Upload a .xlsx, .xlsm, or .csv file (use the template)")

    if not rows:
        raise HTTPException(status_code=400, detail="File is empty")

    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    col_index = {name: header.index(name) for name in TEMPLATE_HEADERS if name in header}

    missing_required = [c for c in ("sku", "name") if c not in col_index]
    if missing_required:
        raise HTTPException(status_code=400, detail=f"Missing required column(s): {', '.join(missing_required)}")

    categories_by_slug = {c.slug: c.id for c in db.query(Category).all()}
    brands_by_slug = {b.slug: b.id for b in db.query(Brand).all()}
    manufacturers_by_name = {m.name.strip().lower(): m for m in db.query(Manufacturer).all()}
    marketers_by_name = {m.name.strip().lower(): m for m in db.query(Marketer).all()}
    # variant_group_key -> variant_group_id, populated as keys are first seen
    # in this import run, so multiple rows sharing a key land in one group
    variant_groups_by_key: dict[str, str] = {}

    created, updated, errors = 0, 0, []

    for row_num, row in enumerate(rows[1:], start=2):
        if row is None or all(v in (None, "") for v in row):
            continue

        def get(col):
            idx = col_index.get(col)
            return row[idx] if idx is not None and idx < len(row) else None

        sku = str(get("sku") or "").strip()
        name = str(get("name") or "").strip()
        if not sku or not name:
            errors.append({"row": row_num, "message": "sku and name are both required"})
            continue
        if sku.upper().startswith("DELETE-THIS-ROW"):
            errors.append({"row": row_num, "message": "Skipped — this is one of the template's example rows, not a real product"})
            continue

        try:
            savepoint = db.begin_nested()
            category_slug = str(get("category_slug") or "").strip()
            brand_slug = str(get("brand_slug") or "").strip()
            category_id = categories_by_slug.get(category_slug) if category_slug else None
            brand_id = brands_by_slug.get(brand_slug) if brand_slug else None
            if category_slug and category_id is None:
                errors.append({"row": row_num, "message": f"Unknown category_slug '{category_slug}' — row still imported, category left blank"})
            if brand_slug and brand_id is None:
                errors.append({"row": row_num, "message": f"Unknown brand_slug '{brand_slug}' — row still imported, brand left blank"})

            # Manufacturer/marketer: looked up by name, auto-created (with no
            # address) if not found — mirrors the "+ Add New" quick-add in
            # the product edit screen. Flagged as a warning so the admin
            # knows to go fill in the address afterward.
            manufacturer_id = None
            manufacturer_name = str(get("manufacturer_name") or "").strip()
            if manufacturer_name:
                existing = manufacturers_by_name.get(manufacturer_name.lower())
                if not existing:
                    existing = Manufacturer(name=manufacturer_name)
                    db.add(existing)
                    db.flush()
                    manufacturers_by_name[manufacturer_name.lower()] = existing
                    errors.append({"row": row_num, "message": f"Created new manufacturer '{manufacturer_name}' with no address — add one in Compliance settings"})
                manufacturer_id = existing.id

            marketer_id = None
            marketer_name = str(get("marketer_name") or "").strip()
            if marketer_name:
                existing = marketers_by_name.get(marketer_name.lower())
                if not existing:
                    existing = Marketer(name=marketer_name)
                    db.add(existing)
                    db.flush()
                    marketers_by_name[marketer_name.lower()] = existing
                    errors.append({"row": row_num, "message": f"Created new marketer '{marketer_name}' with no address — add one in Compliance settings"})
                marketer_id = existing.id

            product = db.query(Product).filter(Product.sku == sku).first()
            is_new = product is None
            if is_new:
                product = Product(sku=sku, slug=_slugify(name) + "-" + sku.lower())
                db.add(product)

            product.name = name
            product.description = str(get("description") or "").strip() or None
            product.image_urls = str(get("image_urls") or "").strip() or None
            product.category_id = category_id
            product.brand_id = brand_id
            product.is_prescription_required = _parse_bool(get("is_prescription_required"))
            product.manufacturer_id = manufacturer_id
            product.marketer_id = marketer_id
            product.country_of_origin = str(get("country_of_origin") or "").strip() or None
            expiry_month = _parse_num(get("compliance_expiry_month"))
            expiry_year = _parse_num(get("compliance_expiry_year"))
            product.expiry_month = int(expiry_month) if expiry_month else None
            product.expiry_year = int(expiry_year) if expiry_year else None
            product.variant_label = str(get("variant_label") or "").strip() or None
            product.batch_number = str(get("batch_number") or "").strip() or None
            product.expiry_date = _parse_date(get("batch_expiry_date"))
            product.rack_place = str(get("rack_place") or "").strip() or None
            product.is_spotlighted = _parse_bool(get("is_spotlighted"))
            spotlight_order = _parse_num(get("spotlight_order"))
            product.spotlight_order = int(spotlight_order) if spotlight_order else 0
            db.flush()

            variant_key = str(get("variant_group_key") or "").strip()
            if variant_key:
                if variant_key not in variant_groups_by_key:
                    variant_groups_by_key[variant_key] = product.variant_group_id or secrets.token_hex(8)
                product.variant_group_id = variant_groups_by_key[variant_key]

            # replace pricing rows for channels present in this row
            channel_fields = {
                PricingChannel.b2c: ("b2c_price", "b2c_mrp", None, "b2c_stock", "b2c_reorder_level"),
                PricingChannel.b2b_normal: ("b2b_normal_price", "b2b_normal_mrp", "b2b_normal_min_qty", "b2b_normal_stock", "b2b_normal_reorder_level"),
                PricingChannel.b2b_advance: ("b2b_advance_price", "b2b_advance_mrp", "b2b_advance_min_qty", "b2b_advance_stock", "b2b_advance_reorder_level"),
                PricingChannel.cnf: ("cnf_price", "cnf_mrp", "cnf_min_qty", "cnf_stock", "cnf_reorder_level"),
            }
            for channel, (price_col, mrp_col, moq_col, stock_col, rol_col) in channel_fields.items():
                price = _parse_num(get(price_col))
                if price is None:
                    continue
                pricing = db.query(ProductPricing).filter(
                    ProductPricing.product_id == product.id, ProductPricing.channel == channel
                ).first()
                if not pricing:
                    pricing = ProductPricing(product_id=product.id, channel=channel)
                    db.add(pricing)
                pricing.price = price
                pricing.mrp = _parse_num(get(mrp_col)) if mrp_col else None
                pricing.min_quantity = int(_parse_num(get(moq_col)) or 1) if moq_col else 1
                pricing.stock = int(_parse_num(get(stock_col)) or 0)
                pricing.reorder_level = int(_parse_num(get(rol_col)) or 0)

            if is_new:
                created += 1
            else:
                updated += 1
            savepoint.commit()

        except Exception as e:
            savepoint.rollback()
            errors.append({"row": row_num, "message": str(e)})
            continue

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}


# ---------- Invoice-based restocking ----------
# Deliberately two separate steps, not one: extract (read-only, just returns
# a best-effort guess) and apply (writes to the DB, only for rows an admin
# has reviewed/corrected and explicitly confirmed). Never auto-applies.

MAX_INVOICE_FILE_SIZE = 15 * 1024 * 1024  # 15MB


@router.post("/extract-invoice")
async def extract_invoice_endpoint(file: UploadFile = File(...), _admin: User = Depends(require_admin)):
    contents = await file.read()
    if len(contents) > MAX_INVOICE_FILE_SIZE:
        raise HTTPException(status_code=400, detail="File is too large (max 15MB)")

    content_type = file.content_type or ""
    try:
        result = extract_invoice(contents, content_type)
    except InvoiceExtractionError as e:
        raise HTTPException(status_code=422, detail=str(e))

    return result


class ApplyInvoiceItem(BaseModel):
    product_id: int
    channel: PricingChannel = PricingChannel.b2c
    quantity_received: int
    batch_number: str | None = None
    expiry_date: str | None = None  # YYYY-MM-DD, applied to the product's batch tracking, not compliance expiry


@router.post("/apply-invoice")
def apply_invoice(items: list[ApplyInvoiceItem], db: Session = Depends(get_db), _admin: User = Depends(require_admin)):
    """
    Adds quantity_received to existing stock (a restock, not an overwrite —
    two invoices for the same product in the same day both count). Never
    touches price or mrp: an invoice's unit cost is what you paid the
    supplier, not what you charge customers, and there's no cost-price field
    in this schema to log it against separately — conflating the two would
    silently change customer-facing prices to purchase cost. Selling price
    stays a manual decision via the product edit screen.
    """
    updated, errors = 0, []
    for item in items:
        pricing = (
            db.query(ProductPricing)
            .filter(ProductPricing.product_id == item.product_id, ProductPricing.channel == item.channel)
            .first()
        )
        if not pricing:
            errors.append({"product_id": item.product_id, "message": f"No {item.channel.value} pricing row exists for this product yet — add pricing first"})
            continue

        pricing.stock += item.quantity_received

        product = db.get(Product, item.product_id)
        if product:
            if item.batch_number:
                product.batch_number = item.batch_number
            if item.expiry_date:
                try:
                    product.expiry_date = _parse_date(item.expiry_date)
                except ValueError:
                    errors.append({"product_id": item.product_id, "message": f"Couldn't parse expiry date '{item.expiry_date}' — stock was still updated"})

        updated += 1

    db.commit()
    return {"updated": updated, "errors": errors}
